"""     All the modules used     """

import csv
import sys
import json
import pprint

from termcolor import colored
from requests import Request, Session
from datetime import datetime

from etherscan import Etherscan
from openpyxl import Workbook, load_workbook
from variable_dic import *
from analyze_data import *
from analyze_data1 import *
from crypto_data import *


"""     Setting up the environment      """

eth = Etherscan(YOUR_API_KEY)


"""     Main Functions       """

def write_daily_account_balance(day, ws_new, name, index_till_25):
    row_index = chr(64 + day)
    cnt = 1

    for row_cells, row_cells1 in zip(ws_new.iter_cols(min_row=day, max_row=day),  ws_new.iter_cols(min_row=1, max_row=1)):
        cnt += 1
        
        for cell, cell1 in zip(row_cells, row_cells1):
            print(cnt)
            
            if cnt < 100:
                cell_value = -1
                
                """     To Avoid Error Api From Etherscan       """
                
                while cell_value == -1:
                    try:
                        cell_value = eth.get_acc_balance_by_token_and_contract_address(name, cell1.value)
                    except json.decoder.JSONDecodeError:
                        print("Json Decoder Error")

                cell.value = round(int(float(cell_value)) / (divided_by[index_till_25]) , 2)


def update_for_all_25_upddates():
    index = 0
    index_till_25 = -1

    for key, value in dic_var.items():
        index = index + 1

        if 'holders' in key:
            index_till_25 = index_till_25 + 1
            
            day_of_the_year = datetime.now().timetuple().tm_yday - 18

            coin_name_address = list(dic_var.items())[index - 2][1]
            

            wb_new = load_workbook(value)
            ws_new = wb_new.active
            
            
            print(value) 

            write_daily_account_balance(day_of_the_year, ws_new, coin_name_address, index_till_25)


            wb_new.save(value)


def main():
    
    update_for_all_25_upddates()

    #analysis()
    #analysis1()


if __name__ == "__main__":
    main()
