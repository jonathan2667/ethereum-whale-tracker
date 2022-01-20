"""
    Analysis

    Try to analyze crypto data:
    
        Global Analysis of the market day:
            A. Information extracted based on a short analysis:
                1.THE TOP HOLDERS VOLUME
                2.CVD OF TOP HOLDERS
                3.NUMBERS ACCOUNT THAT ADDED
                4.NUMBERS ACCOUNT THAT SOLD
                5.NOUMBERS ACCOUNT THAT NULL
                
            B. Information extracted from coinmarketcap API, from cytpto_data.py:
                1.VOLUME 24h
                2.CURRENT PRICE
                3.PERCENT CHANGE
                4.OVERALL VOLUME 24 H
        
    Codes need to be verified for more accurate information!
        
"""

"""     All the modules used     """

from openpyxl import Workbook, load_workbook
from datetime import datetime

from main1 import *
from crypto_data import *
from variable_dic import *


"""     Main Functions       """


def get(value, index_till_25):
    return round((int(float(value))/divided_by[index_till_25]), 2)

def travel_over_data(ws_new, index_till_25):
    day = datetime.now().timetuple().tm_yday - 10

    portfolios_that_added = 0
    portfolios_that_sold = 0
    portfolios_no_action = 0
    volume_total_top_holders = 0
    cvd_top_holders = 0

    row_index = chr(64 + day)

    cnt = 1

    for row_cells, row_cells1 in zip(ws_new.iter_cols(min_row=day, max_row=day), ws_new.iter_cols(min_row=day-1, max_row=day-1)):
        cnt += 1
        for cell, cell1 in zip(row_cells, row_cells1):
            if cnt < 100:
                if cell.value > cell1.value :
                    portfolios_that_added = portfolios_that_added + 1
                elif cell.value < cell1.value:
                    portfolios_that_sold = portfolios_that_sold + 1
                else:
                    portfolios_no_action = portfolios_no_action + 1
                volume_total_top_holders = volume_total_top_holders + abs(cell.value - cell1.value)
                cvd_top_holders = cvd_top_holders + cell.value - cell1.value

    volume_total_top_holders = round(volume_total_top_holders, 2)

    return volume_total_top_holders, round(cvd_top_holders, 2),  portfolios_that_added, portfolios_that_sold, portfolios_no_action


def analysis():

    index = 0
    index_till_25 = -1

    for key, value in dic_var.items():
        index = index + 1

        if 'holders' in key:
            color = ''

            index_till_25 = index_till_25 + 1

            coin_name = list(dic_var.items())[index - 2][0]
            print(coin_name)

            wb_new = load_workbook(value)
            ws_new = wb_new.active

            volume_top_holders, cvd_top_holders, portfolios_that_added, portfolios_that_sold, portfolios_no_action = travel_over_data(ws_new, index_till_25)
            curr_price, percent_change_24h, overall_volume_change_24h, volume_24h = crypto(list_symbol[index_till_25 - 1])

            if portfolios_that_sold > portfolios_that_added:
                color = 'red'
            elif portfolios_that_sold < portfolios_that_added:
                color = 'green'
            else:
                color = 'white'

            print(list_symbol[index_till_25 - 1])
            print(colored(" THE TOP HOLDERS VOLUME IS : ", color), colored(volume_top_holders * curr_price, color), colored(" CVD OF TOP HOLDERS: ", color), colored(cvd_top_holders, color), colored(" ADDED: ", color), colored(portfolios_that_added, color), colored("   SOLD: ", color), colored(portfolios_that_sold, color), colored( "    NULL:", color), colored(portfolios_no_action, color))
            print(colored(" VOLUME 24h: ", color), colored(volume_24h, color) , colored("   CURRENT PRICE: ", color), colored(curr_price, color), colored("  PERCENT CHANGE: ", color), colored(percent_change_24h, color), colored('%', color), colored("   OVERALL VOLUME 24 H: ", color) , colored(overall_volume_change_24h, color), colored("%", color))

            print('')


