"""     All the modules used     """

import csv
import sys
import pprint
import json

from etherscan import Etherscan
from openpyxl import Workbook, load_workbook

from variable import *
from variable_dic import *
from datetime import datetime


"""     Setting up the environment      """

eth = Etherscan(YOUR_API_KEY)

wb_coin_addresses = load_workbook('CoinAdresses.xlsx')
ws_coin_adresses = wb_coin_addresses.active


"""     Main Functions       """

def get_holders_from_csv_file(file_holders_etherscan, index_till_25):
    with open(file_holders_etherscan) as csv_file:
        csv_reader = csv.reader(csv_file)

        for line in csv_reader:
            try:
                holders[line[0]] = round(int(float(line[1])), 2)
            except ValueError:
                do_nothing = 1

        sort_holders = sorted(holders.items(), key = lambda x:x[1], reverse= True)

        return sort_holders


def write_holders_in_addresses_file(holders_f, ind, coin_name):
    row_index = chr(64 + ind)
    ws_coin_adresses[row_index + '1'].value = coin_name

    for i in range (99):
        cell = row_index + str(i + 2)
        ws_coin_adresses[cell] = holders_f[i][0]


def write_holders_data_in_day_tracker(ind, ws_new):
    row_index = chr(64 + ind)
    cnt = 1

    for row_cells in ws_new.iter_cols(min_row=1, max_row=1, max_col=99):
        cnt += 1

        for cell in row_cells:
            cell.value = str(ws_coin_adresses[row_index + str(cnt)].value)


def write_daily_account_balance(day, ws_new, coin_name_address, index_till_25 ):
    row_index = chr(64 + day)
    cnt = 1

    for row_cells, row_cells1 in zip(ws_new.iter_cols(min_row=day, max_row=day),  ws_new.iter_cols(min_row=1, max_row=1)):
        cnt += 1

        for cell, cell1 in zip(row_cells, row_cells1):
            print(cnt)

            if cnt < 100:
                cell_value = -1
                
                """     To Avoid Error Api From Etherscan       """
                
                while cell_value == -1:
                    try:
                        cell_value = eth.get_acc_balance_by_token_and_contract_address(coin_name_address, cell1.value)
                    except json.decoder.JSONDecodeError:
                        print("Json Decoder Error")

                cell.value = round(int(float(cell_value)) / (divided_by[index_till_25]) , 2)


def main():
    index = 0
    index_till_25 = -1

    for key, value in dic_var.items():

        index = index + 1

        if 'holders' in key:

            """     Global variables for every file      """

            index_till_25 = index_till_25 + 1

            day_of_the_year = datetime.now().timetuple().tm_yday - 18

            coin_name_address = list(dic_var.items())[index - 2][1]
            coin_name = list(dic_var.items())[index - 2][0]

            coin_file = list(dic_var.items())[index - 3][1]

            excel_uptime_etherscan_name = list(dic_var.items())[index - 1][1]


            wb_new = load_workbook(excel_uptime_etherscan_name)
            ws_new = wb_new.active


            """     Main Functions Here     """

            holders_f = get_holders_from_csv_file(coin_file, index_till_25)                             #takes from the exported holders files from etherscan.io of a coin the first 100 holcers by quantity               #change here

            write_holders_in_addresses_file(holders_f, index_till_25 + 1, coin_name)                    #writes the top 100 holders of the coin in the coinaddresses.xlsx           # CANGE HERE!!!!! index and name

            write_holders_data_in_day_tracker(index_till_25 + 1, ws_new)                                #write the in the first row the top 100 holders

            write_daily_account_balance(day_of_the_year, ws_new, coin_name_address, index_till_25)      #emits query based on the holder and writes it down based on the DAY!!!!


            wb_new.save(excel_uptime_etherscan_name)
            wb_coin_addresses.save('CoinAdresses.xlsx')


if __name__ == "__main__":
    main()



