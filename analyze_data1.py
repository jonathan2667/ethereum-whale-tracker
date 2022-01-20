"""
    Analysis 1

    Try to analyze crypto data:

        A. Per HOLDER:
            1. TRADE SIZE INDICAtOR for different time frames
            2. CVD for different time frames

        B. For top 100 HOLDERS:
            1. TRADE SIZE INDICAtOR for different time frames
            2. CVD for different time frames

        How:

        - GET THE DATA for each tool
        - PROCESS THE DATA with PANDA or MATPLOTLIB
        _ STORE THE processed data (PNG, etc)
"""


"""     All the modules used     """

import sys

from openpyxl import Workbook, load_workbook

from datetime import datetime
from main1 import *
from variable_dic import *
from crypto_data import *


def get(value, index_till_25):
    return round((int(float(value))/divided_by[index_till_25]), 2)

def trade_size_indicator(period, ws_new, index_till_25):
    period = period - 1
    day = datetime.now().timetuple().tm_yday - 10 - 1 #erase the -1 at the end, just for backtesting purposes

    cnt = 1


    for row_cells in ws_new.iter_cols(min_row=day - period, max_row=day):
        cnt = cnt + 1
        initial_balance = 0
        total_coin_transactions = 0

        for cell in row_cells:
            if cnt < 100:
                # if initial_balance == 0:
                #     initial_balance = get(cell.value)
                # else:
                #     total_coin_transactions = total_coin_transactions + (get(cell.value) - initial_balance)
                #     initial_balance = get(cell.value)
                print(get(cell.value, index_till_25))
        print()


def analysis1() :
    index = 0
    index_till_25 = -1

    for key, value in dic_var.items():
        index = index + 1

        if 'holders' in key:
            index_till_25 = index_till_25 + 1

            wb_new = load_workbook(value)
            ws_new = wb_new.active

            print(list(dic_var.items())[index - 2][0])

            time_frames = [1]

            for period in time_frames:
                trade_size_indicator(period, ws_new, index_till_25)
